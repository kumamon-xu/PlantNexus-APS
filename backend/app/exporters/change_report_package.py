"""Deterministic P4 ChangeReport package layered on a verified P3 package."""

from __future__ import annotations

from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from hashlib import sha256
import io
import json
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any, Never, cast
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.execution_contracts import (
    canonical_contract_bytes,
    contract_fingerprint,
    export_manifest_fingerprint,
    require_p4_document,
)
from app.domain.workspace_contracts import require_workspace_document
from app.exporters.standard_package import (
    MAX_STANDARD_EXPORT_FILE_BYTES,
    MAX_STANDARD_EXPORT_MANIFEST_BYTES,
    MAX_STANDARD_EXPORT_PACKAGE_BYTES,
    StandardExportError,
    StandardExportErrorCode,
    StandardExportPackage,
    _canonicalize_xlsx,
    _csv_rows,
    _safe_cell,
    storage_reference_for,
    verify_standard_export_package,
)


EXPORT_MANIFEST_VERSION = "export-manifest.v3"
EXPORT_JOB_VERSION = "export-job.v3"
EXPORT_SCHEMA_SET_VERSION = "2.8.0"
EXPORT_PACKAGE_PROFILE = "p4-dynamic-replan-export.v1"
CHANGE_REPORT_READ_PROFILE = "change-report.v1"
XLSX_PROFILE_VERSION = "xlsx-safe-deterministic.v1"
CSV_DIALECT_VERSION = "rfc4180-lf.v1"
ZIP_PROFILE_VERSION = "zip-deterministic.v1"

_PAYLOAD_ROLES = {
    "change_report.json": "CHANGE_REPORT",
    "import_quality_report.json": "IMPORT_QUALITY_REPORT",
    "kpi.json": "KPI",
    "order_summary.csv": "ORDER_SUMMARY",
    "planning_solution.json": "VALIDATED_PLANNING_SOLUTION",
    "publication_result.json": "PUBLICATION_RESULT",
    "resource_load.csv": "RESOURCE_LOAD",
    "scenario_manifest.json": "P2_CORRECTNESS_SCENARIO_MANIFEST",
    "schedule_operations.csv": "SCHEDULE_OPERATIONS",
    "schedule_version.json": "PUBLISHED_SCHEDULE_VERSION",
    "solver_report.json": "SOLVER_REPORT",
    "standard_package.xlsx": "STANDARD_WORKBOOK",
    "validation_report.json": "VALIDATION_REPORT",
}
_P3_REUSED_PATHS = (
    "planning_solution.json",
    "schedule_operations.csv",
    "order_summary.csv",
    "resource_load.csv",
    "import_quality_report.json",
    "scenario_manifest.json",
)
_DEFERRED = [
    {"path": "benchmark_report.json", "status": "NOT_APPLICABLE_STANDARD_EXPORT"}
]
_WORKBOOK_SHEETS = (
    ("Schedule Operations", "schedule_operations.csv"),
    ("Order Summary", "order_summary.csv"),
    ("Resource Load", "resource_load.csv"),
)
_FINGERPRINT = re.compile(r"sha256:[0-9a-f]{64}")

type JsonObject = dict[str, Any]


@dataclass(frozen=True, slots=True)
class ChangeReportExportPackage:
    package_id: str
    manifest_fingerprint: str
    storage_reference: str
    _files: tuple[tuple[str, bytes], ...]

    @property
    def files(self) -> dict[str, bytes]:
        return dict(self._files)

    @property
    def manifest(self) -> JsonObject:
        return cast(JsonObject, json.loads(self.files["manifest.json"]))


def _reject(code: StandardExportErrorCode, field: str) -> Never:
    raise StandardExportError(code, field=field)


def _fingerprint(value: bytes) -> str:
    return f"sha256:{sha256(value).hexdigest()}"


def _json_bytes(value: Mapping[str, object]) -> bytes:
    return canonical_contract_bytes(value)


def _mapping(value: object, field: str) -> Mapping[str, object]:
    if not isinstance(value, Mapping):
        _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
    return cast(Mapping[str, object], value)


def _sequence(value: object, field: str) -> Sequence[object]:
    if not isinstance(value, list):
        _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
    return cast(Sequence[object], value)


def _report_reference(report: Mapping[str, object]) -> dict[str, object]:
    return {
        "change_report_version": "change-report.v1",
        "report_id": report["report_id"],
        "report_fingerprint": report["report_fingerprint"],
    }


def _schedule_reference(schedule: Mapping[str, object]) -> dict[str, object]:
    return {
        "schedule_version_version": "schedule-version.v2",
        "schedule_version_id": schedule["schedule_version_id"],
        "state": "PUBLISHED",
        "content_fingerprint": schedule["content_fingerprint"],
    }


def _validation_fingerprint(validation: Mapping[str, object]) -> str:
    return contract_fingerprint(
        _mapping(validation.get("formal_validation"), "validation_report.formal_validation")
    )


def _change_report_rows(report: Mapping[str, object]) -> list[list[str]]:
    rows = [[
        "operation_id",
        "classification",
        "base_resource_id",
        "new_resource_id",
        "start_shift_seconds",
        "duration_delta_seconds",
        "reason_codes",
    ]]
    for raw_operation in _sequence(report.get("operations"), "change_report.operations"):
        operation = _mapping(raw_operation, "change_report.operations[]")
        base = operation.get("base_assignment")
        new = operation.get("new_assignment")
        deltas = _mapping(operation.get("deltas"), "change_report.operations[].deltas")
        reasons = _sequence(operation.get("reasons"), "change_report.operations[].reasons")
        rows.append(
            [
                cast(str, operation["operation_id"]),
                cast(str, operation["classification"]),
                "" if base is None else cast(str, _mapping(base, "base_assignment")["resource_id"]),
                "" if new is None else cast(str, _mapping(new, "new_assignment")["resource_id"]),
                str(deltas["start_shift_seconds"]),
                str(deltas["duration_delta_seconds"]),
                "|".join(
                    cast(str, _mapping(reason, "reasons[]")["reason_code"])
                    for reason in reasons
                ),
            ]
        )
    return rows


def _workbook_bytes(
    payloads: Mapping[str, bytes],
    report: Mapping[str, object],
    metadata: Mapping[str, object],
) -> bytes:
    workbook = Workbook()
    workbook.properties.creator = "PlantNexus APS"
    workbook.properties.lastModifiedBy = "PlantNexus APS"
    workbook.properties.created = datetime(1980, 1, 1, tzinfo=UTC)
    workbook.properties.modified = datetime(1980, 1, 1, tzinfo=UTC)
    workbook.properties.title = "P4 Dynamic Replan Export"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    first = True
    for title, path in _WORKBOOK_SHEETS:
        sheet = cast(Worksheet, workbook.active) if first else workbook.create_sheet()
        first = False
        sheet.title = title
        for row in _csv_rows(payloads[path], path):
            sheet.append([_safe_cell(value) for value in row])
    sheet = workbook.create_sheet("Change Report")
    for row in _change_report_rows(report):
        sheet.append([_safe_cell(value) for value in row])
    sheet = workbook.create_sheet("Metadata")
    sheet.append(["key", "value"])
    for key, value in sorted(metadata.items()):
        rendered = json.dumps(
            value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
        sheet.append([_safe_cell(str(key)), _safe_cell(rendered)])
    stream = io.BytesIO()
    workbook.save(stream)
    rendered = _canonicalize_xlsx(stream.getvalue())
    _verify_xlsx(rendered)
    return rendered


def _verify_xlsx(value: bytes) -> int:
    expected_sheets = [title for title, _ in _WORKBOOK_SHEETS] + [
        "Change Report",
        "Metadata",
    ]
    try:
        with ZipFile(io.BytesIO(value), "r") as archive:
            names = set(archive.namelist())
            forbidden = (
                "vbaProject.bin",
                "externalLinks/",
                "embeddings/",
                "connections.xml",
            )
            if any(any(token in name for token in forbidden) for name in names):
                _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
            if any(
                re.search(rb"<f(?:[ >])", archive.read(name))
                for name in names
                if name.endswith(".xml")
            ):
                _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
        workbook = load_workbook(io.BytesIO(value), read_only=False, data_only=False)
        if workbook.sheetnames != expected_sheets:
            _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
        if any(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        ):
            _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
        return len(workbook.sheetnames)
    except StandardExportError:
        raise
    except Exception:
        _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")


def _file_record(path: str, content: bytes) -> JsonObject:
    is_csv = path.endswith(".csv")
    is_xlsx = path.endswith(".xlsx")
    return {
        "path": path,
        "role": _PAYLOAD_ROLES[path],
        "media_type": (
            "text/csv; charset=utf-8"
            if is_csv
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if is_xlsx
            else "application/json"
        ),
        "sha256": _fingerprint(content),
        "size_bytes": len(content),
        "row_count": len(_csv_rows(content, path)) - 1 if is_csv else None,
        "sheet_count": _verify_xlsx(content) if is_xlsx else None,
    }


def _validate_inputs(
    *,
    p3_package: StandardExportPackage,
    schedule_version: Mapping[str, object],
    publication_result: Mapping[str, object],
    export_job: Mapping[str, object],
    change_report: Mapping[str, object],
    solver_report: Mapping[str, object],
    validation_report: Mapping[str, object],
    kpi: Mapping[str, object],
) -> None:
    verify_standard_export_package(p3_package)
    try:
        versions = {
            require_p4_document(schedule_version),
            require_p4_document(export_job),
            require_p4_document(change_report),
            require_p4_document(solver_report),
        }
    except (TypeError, ValueError) as error:
        raise StandardExportError(
            StandardExportErrorCode.INVALID_PACKAGE,
            field=cast(str, getattr(error, "field", "p4_document")),
        ) from error
    if versions != {
        "schedule-version.v2",
        "export-job.v3",
        "change-report.v1",
        "solver-report.v2",
    }:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "p4_document_versions")
    if require_workspace_document(publication_result) != "publication-result.v1":
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "publication_result")
    if (
        schedule_version.get("state") != "PUBLISHED"
        or export_job.get("state") != "EXPORTING"
        or not isinstance(export_job.get("attempt"), int)
        or cast(int, export_job["attempt"]) < 1
        or solver_report.get("solver_status") not in {"OPTIMAL", "FEASIBLE"}
        or validation_report.get("status") != "PASS"
        or validation_report.get("hard_violation_count") != 0
        or kpi.get("kpi_version") != "kpi.v2"
    ):
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "state/validation")
    if (
        schedule_version.get("data_plane") != "SIMULATION"
        or export_job.get("data_plane") != "SIMULATION"
        or schedule_version.get("synthetic") is not True
        or export_job.get("synthetic") is not True
        or export_job.get("target") != "SIMULATION_INTERNAL"
        or export_job.get("package_profile") != EXPORT_PACKAGE_PROFILE
    ):
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "plane/target/synthetic")

    schedule_reference = _schedule_reference(schedule_version)
    report_reference = _report_reference(change_report)
    published = _mapping(
        publication_result.get("published_version"),
        "publication_result.published_version",
    )
    report_new = _mapping(
        change_report.get("new_schedule_version"),
        "change_report.new_schedule_version",
    )
    report_lineage = _mapping(change_report.get("lineage"), "change_report.lineage")
    schedule_lineage = _mapping(schedule_version.get("lineage"), "schedule_version.lineage")
    if (
        export_job.get("schedule_version") != schedule_reference
        or export_job.get("change_report") != report_reference
        or published.get("schedule_version_id") != schedule_reference["schedule_version_id"]
        or published.get("content_fingerprint") != schedule_reference["content_fingerprint"]
        or report_new.get("schedule_version_id") != schedule_reference["schedule_version_id"]
        or report_new.get("content_fingerprint") != schedule_reference["content_fingerprint"]
        or schedule_lineage.get("change_report") != report_reference
        or schedule_lineage.get("replan_request") != report_lineage.get("replan_request")
        or schedule_lineage.get("planning_run_id") != report_lineage.get("planning_run_id")
    ):
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "schedule/report/job/publication")

    solver_reference = _mapping(
        report_lineage.get("solver_report"), "change_report.lineage.solver_report"
    )
    validation_reference = _mapping(
        report_lineage.get("validation_report"),
        "change_report.lineage.validation_report",
    )
    after_kpi = _mapping(change_report.get("after_kpi"), "change_report.after_kpi")
    validation_fingerprint = _validation_fingerprint(validation_report)
    if (
        solver_reference.get("artifact_id") != solver_report.get("report_id")
        or solver_reference.get("fingerprint") != solver_report.get("report_fingerprint")
        or validation_reference.get("artifact_id")
        != "validation-report-"
        + validation_fingerprint.removeprefix("sha256:")
        or validation_reference.get("fingerprint") != validation_fingerprint
        or after_kpi.get("artifact_id") != kpi.get("kpi_id")
        or after_kpi.get("fingerprint") != contract_fingerprint(kpi)
        or schedule_lineage.get("solver_report") != solver_reference
        or schedule_lineage.get("validation_report") != validation_reference
        or schedule_lineage.get("kpi") != after_kpi
    ):
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "artifact_lineage")

    p3_schedule = cast(
        dict[str, object], json.loads(p3_package.files["schedule_version.json"])
    )
    p3_solution = cast(
        dict[str, object], json.loads(p3_package.files["planning_solution.json"])
    )
    schedule_content = _mapping(schedule_version.get("content"), "schedule_version.content")
    p3_content = _mapping(p3_schedule.get("content"), "p3.schedule_version.content")
    if (
        p3_content != schedule_content
        or p3_schedule.get("content_fingerprint")
        != schedule_version.get("content_fingerprint")
        or p3_solution.get("assignments") != schedule_content.get("assignments")
        or p3_package.manifest.get("target") != "SIMULATION_INTERNAL"
        or p3_package.manifest.get("publishable") is not False
    ):
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "p3_package/content")


def build_change_report_export_package(
    *,
    p3_package: StandardExportPackage,
    schedule_version: Mapping[str, object],
    publication_result: Mapping[str, object],
    export_job: Mapping[str, object],
    change_report: Mapping[str, object],
    solver_report: Mapping[str, object],
    validation_report: Mapping[str, object],
    kpi: Mapping[str, object],
    create_audit_event_id: str,
    attempt_audit_event_id: str,
    completion_audit_event_id: str,
    correlation_id: str,
    generated_at_utc: str,
) -> ChangeReportExportPackage:
    """Add exact P4 artifacts while retaining a verified compatible P3 package."""

    _validate_inputs(
        p3_package=p3_package,
        schedule_version=schedule_version,
        publication_result=publication_result,
        export_job=export_job,
        change_report=change_report,
        solver_report=solver_report,
        validation_report=validation_report,
        kpi=kpi,
    )
    p3_files = p3_package.files
    payloads = {path: p3_files[path] for path in _P3_REUSED_PATHS}
    payloads.update(
        {
            "schedule_version.json": _json_bytes(schedule_version),
            "publication_result.json": _json_bytes(publication_result),
            "change_report.json": _json_bytes(change_report),
            "solver_report.json": _json_bytes(solver_report),
            "validation_report.json": _json_bytes(validation_report),
            "kpi.json": _json_bytes(kpi),
        }
    )
    payloads["standard_package.xlsx"] = _workbook_bytes(
        payloads,
        change_report,
        {
            "export_job_id": export_job["export_job_id"],
            "attempt": export_job["attempt"],
            "schedule_version_id": schedule_version["schedule_version_id"],
            "content_fingerprint": schedule_version["content_fingerprint"],
            "change_report_id": change_report["report_id"],
            "change_report_fingerprint": change_report["report_fingerprint"],
            "publication_id": publication_result["publication_id"],
            "target": "SIMULATION_INTERNAL",
            "profile": EXPORT_PACKAGE_PROFILE,
        },
    )
    file_records = [_file_record(path, payloads[path]) for path in sorted(payloads)]
    p3_counts = _mapping(p3_package.manifest.get("entity_counts"), "p3.entity_counts")
    manifest_basis: JsonObject = {
        "export_manifest_version": EXPORT_MANIFEST_VERSION,
        "schema_set_version": EXPORT_SCHEMA_SET_VERSION,
        "package_profile": EXPORT_PACKAGE_PROFILE,
        "canonicalization_version": "canonical-json.v1",
        "csv_dialect_version": CSV_DIALECT_VERSION,
        "xlsx_profile_version": XLSX_PROFILE_VERSION,
        "generated_at_utc": generated_at_utc,
        "publishable": False,
        "target": "SIMULATION_INTERNAL",
        "schedule_version": _schedule_reference(schedule_version),
        "publication": {
            "publication_result_version": publication_result["publication_result_version"],
            "publication_id": publication_result["publication_id"],
            "result_fingerprint": publication_result["result_fingerprint"],
            "published_at_utc": publication_result["published_at_utc"],
            "audit_event_id": publication_result["audit_event_id"],
        },
        "export_job": {
            "export_job_version": EXPORT_JOB_VERSION,
            "export_job_id": export_job["export_job_id"],
            "attempt": export_job["attempt"],
            "state_at_materialization": "EXPORTING",
        },
        "audit_lineage": {
            "create_audit_event_id": create_audit_event_id,
            "attempt_audit_event_id": attempt_audit_event_id,
            "completion_audit_event_id": completion_audit_event_id,
            "correlation_id": correlation_id,
        },
        "p3_package": {
            "export_manifest_version": "export-manifest.v2",
            "package_profile": "p3-standard-export.v1",
            "package_id": p3_package.package_id,
            "manifest_fingerprint": p3_package.manifest_fingerprint,
        },
        "change_report": _report_reference(change_report),
        "entity_counts": {
            "snapshot_operation_count": p3_counts["snapshot_operation_count"],
            "problem_operation_count": p3_counts["problem_operation_count"],
            "assignment_count": p3_counts["assignment_count"],
            "change_operation_count": change_report["operation_universe_count"],
            "demand_count": p3_counts["demand_count"],
            "resource_count": p3_counts["resource_count"],
        },
        "file_count": len(file_records),
        "files": file_records,
        "deferred_artifacts": _DEFERRED,
        "state_boundary": {
            "schedule_version": "PUBLISHED",
            "publication": "COMPLETED",
            "export_job_at_materialization": "EXPORTING",
            "external_transfer": "NOT_STARTED",
            "production": "NOT_AUTHORIZED",
        },
        "synthetic": True,
        "synthetic_provenance": schedule_version["synthetic_provenance"],
    }
    semantic_fingerprint = export_manifest_fingerprint(manifest_basis)
    package_id = "export-package-" + semantic_fingerprint.removeprefix("sha256:")
    manifest = {
        "package_id": package_id,
        "manifest_fingerprint": semantic_fingerprint,
        **manifest_basis,
    }
    manifest_bytes = _json_bytes(manifest)
    package = ChangeReportExportPackage(
        package_id=package_id,
        manifest_fingerprint=semantic_fingerprint,
        storage_reference=storage_reference_for(package_id),
        _files=tuple(sorted({"manifest.json": manifest_bytes, **payloads}.items())),
    )
    verify_change_report_export_package(package)
    return package


def verify_change_report_export_package(package: ChangeReportExportPackage) -> None:
    try:
        files = package.files
        if set(files) != {"manifest.json", *_PAYLOAD_ROLES}:
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "files")
        json_paths = sorted(path for path in files if path.endswith(".json"))
        documents = {
            path: cast(JsonObject, json.loads(files[path])) for path in json_paths
        }
        if any(_json_bytes(documents[path]) != files[path] for path in json_paths):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "canonical_json")
        manifest = documents["manifest.json"]
        schedule = documents["schedule_version.json"]
        report = documents["change_report.json"]
        solver = documents["solver_report.json"]
        validation = documents["validation_report.json"]
        kpi = documents["kpi.json"]
        publication = documents["publication_result.json"]
        if require_p4_document(manifest) != EXPORT_MANIFEST_VERSION:
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "manifest")
        if (
            require_p4_document(schedule) != "schedule-version.v2"
            or require_p4_document(report) != "change-report.v1"
            or require_p4_document(solver) != "solver-report.v2"
            or require_workspace_document(publication) != "publication-result.v1"
        ):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "documents")
        if (
            manifest.get("package_id") != package.package_id
            or manifest.get("manifest_fingerprint") != package.manifest_fingerprint
            or export_manifest_fingerprint(manifest) != package.manifest_fingerprint
            or package.storage_reference != storage_reference_for(package.package_id)
        ):
            _reject(StandardExportErrorCode.HASH_MISMATCH, "manifest_identity")
        records = cast(list[Mapping[str, object]], manifest["files"])
        if [record.get("path") for record in records] != sorted(_PAYLOAD_ROLES):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "manifest.files")
        for record in records:
            path = cast(str, record["path"])
            content = files[path]
            if (
                record.get("role") != _PAYLOAD_ROLES[path]
                or record.get("sha256") != _fingerprint(content)
                or record.get("size_bytes") != len(content)
            ):
                _reject(StandardExportErrorCode.HASH_MISMATCH, path)
            if path.endswith(".csv") and record.get("row_count") != (
                len(_csv_rows(content, path)) - 1
            ):
                _reject(StandardExportErrorCode.HASH_MISMATCH, path)
            if path.endswith(".xlsx") and record.get("sheet_count") != _verify_xlsx(content):
                _reject(StandardExportErrorCode.UNSAFE_XLSX, path)
        manifest_schedule = _mapping(manifest.get("schedule_version"), "manifest.schedule_version")
        manifest_report = _mapping(manifest.get("change_report"), "manifest.change_report")
        schedule_lineage = _mapping(schedule.get("lineage"), "schedule.lineage")
        report_lineage = _mapping(report.get("lineage"), "report.lineage")
        if (
            manifest_schedule != _schedule_reference(schedule)
            or manifest_report != _report_reference(report)
            or schedule_lineage.get("change_report") != _report_reference(report)
            or _mapping(report.get("new_schedule_version"), "report.new_schedule_version").get("schedule_version_id")
            != schedule.get("schedule_version_id")
            or _mapping(report.get("new_schedule_version"), "report.new_schedule_version").get("content_fingerprint")
            != schedule.get("content_fingerprint")
            or report_lineage.get("solver_report") != schedule_lineage.get("solver_report")
            or report_lineage.get("validation_report") != schedule_lineage.get("validation_report")
        ):
            _reject(StandardExportErrorCode.MIXED_LINEAGE, "manifest/report/schedule")
        solver_reference = _mapping(report_lineage.get("solver_report"), "report.solver_report")
        validation_reference = _mapping(
            report_lineage.get("validation_report"), "report.validation_report"
        )
        validation_fingerprint = _validation_fingerprint(validation)
        after_kpi = _mapping(report.get("after_kpi"), "report.after_kpi")
        planning_solution = documents["planning_solution.json"]
        schedule_content = _mapping(schedule.get("content"), "schedule.content")
        if (
            solver.get("solver_status") not in {"OPTIMAL", "FEASIBLE"}
            or solver_reference.get("artifact_id") != solver.get("report_id")
            or solver_reference.get("fingerprint") != solver.get("report_fingerprint")
            or validation.get("status") != "PASS"
            or validation.get("hard_violation_count") != 0
            or validation_reference.get("artifact_id")
            != "validation-report-"
            + validation_fingerprint.removeprefix("sha256:")
            or validation_reference.get("fingerprint") != validation_fingerprint
            or after_kpi.get("artifact_id") != kpi.get("kpi_id")
            or after_kpi.get("fingerprint") != contract_fingerprint(kpi)
            or planning_solution.get("assignments") != schedule_content.get("assignments")
        ):
            _reject(StandardExportErrorCode.MIXED_LINEAGE, "payload_lineage")
        if manifest.get("publishable") is not False or manifest.get("target") != "SIMULATION_INTERNAL":
            _reject(StandardExportErrorCode.MIXED_LINEAGE, "publishable/target")
    except StandardExportError:
        raise
    except Exception:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "package")


def _bounded_file_bytes(path: Path, *, maximum: int, field: str) -> bytes:
    try:
        if path.is_symlink() or not path.is_file():
            _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
        size = path.stat().st_size
        if size < 1 or size > maximum:
            _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
        value = path.read_bytes()
        if len(value) != size:
            _reject(StandardExportErrorCode.IO_ERROR, field)
        return value
    except StandardExportError:
        raise
    except OSError as error:
        raise StandardExportError(StandardExportErrorCode.IO_ERROR, field=field) from error


def load_change_report_export_package(source: Path) -> ChangeReportExportPackage:
    original = source
    try:
        if original.is_symlink():
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "source")
        source = original.resolve(strict=True)
        if not source.is_dir():
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "source")
        children = list(source.iterdir())
        expected_names = {"manifest.json", *_PAYLOAD_ROLES}
        if {path.name for path in children} != expected_names or any(
            path.is_symlink() or not path.is_file() for path in children
        ):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "source.files")
        manifest_bytes = _bounded_file_bytes(
            source / "manifest.json",
            maximum=MAX_STANDARD_EXPORT_MANIFEST_BYTES,
            field="manifest.json",
        )
        manifest = cast(JsonObject, json.loads(manifest_bytes))
        package_id = manifest.get("package_id")
        manifest_fingerprint = manifest.get("manifest_fingerprint")
        if (
            not isinstance(package_id, str)
            or not isinstance(manifest_fingerprint, str)
            or _FINGERPRINT.fullmatch(manifest_fingerprint) is None
        ):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "manifest.identity")
        files: dict[str, bytes] = {"manifest.json": manifest_bytes}
        total_bytes = len(manifest_bytes)
        for name in sorted(_PAYLOAD_ROLES):
            value = _bounded_file_bytes(
                source / name,
                maximum=MAX_STANDARD_EXPORT_FILE_BYTES,
                field=name,
            )
            total_bytes += len(value)
            if total_bytes > MAX_STANDARD_EXPORT_PACKAGE_BYTES:
                _reject(StandardExportErrorCode.INVALID_PACKAGE, "package.size_bytes")
            files[name] = value
        package = ChangeReportExportPackage(
            package_id=package_id,
            manifest_fingerprint=manifest_fingerprint,
            storage_reference=storage_reference_for(package_id),
            _files=tuple(sorted(files.items())),
        )
        verify_change_report_export_package(package)
        return package
    except StandardExportError:
        raise
    except (OSError, json.JSONDecodeError, UnicodeError) as error:
        raise StandardExportError(
            StandardExportErrorCode.INVALID_PACKAGE, field="source"
        ) from error


def archive_change_report_export_package(package: ChangeReportExportPackage) -> bytes:
    verify_change_report_export_package(package)
    if sum(len(value) for value in package.files.values()) > MAX_STANDARD_EXPORT_PACKAGE_BYTES:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "package.size_bytes")
    stream = io.BytesIO()
    try:
        ordered_names = [
            *sorted(name for name in package.files if name != "manifest.json"),
            "manifest.json",
        ]
        with ZipFile(stream, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            for name in ordered_names:
                info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = ZIP_DEFLATED
                info.create_system = 0
                info.external_attr = 0
                archive.writestr(info, package.files[name])
        rendered = stream.getvalue()
        with ZipFile(io.BytesIO(rendered), "r") as replay:
            if replay.namelist() != ordered_names or any(
                replay.read(name) != package.files[name] for name in ordered_names
            ):
                _reject(StandardExportErrorCode.HASH_MISMATCH, "archive")
        return rendered
    except StandardExportError:
        raise
    except Exception as error:
        raise StandardExportError(StandardExportErrorCode.IO_ERROR, field="archive") from error


def change_report_export_bytes_fingerprint(value: bytes) -> str:
    return _fingerprint(value)


def _directory_matches(destination: Path, package: ChangeReportExportPackage) -> bool:
    if not destination.is_dir():
        return False
    children = list(destination.iterdir())
    return (
        all(path.is_file() for path in children)
        and {path.name: path.read_bytes() for path in children} == package.files
    )


def _write_file(path: Path, value: bytes) -> None:
    path.write_bytes(value)


def write_change_report_export_package(
    package: ChangeReportExportPackage,
    destination: Path,
    *,
    file_writer: Callable[[Path, bytes], None] = _write_file,
) -> Path:
    verify_change_report_export_package(package)
    destination = destination.resolve()
    parent = destination.parent
    temporary: Path | None = None
    try:
        if not parent.is_dir():
            _reject(StandardExportErrorCode.IO_ERROR, "destination.parent")
        if destination.exists():
            if _directory_matches(destination, package):
                return destination
            _reject(StandardExportErrorCode.DESTINATION_CONFLICT, "destination")
        temporary = Path(
            tempfile.mkdtemp(prefix=f".{destination.name}.tmp-", dir=parent)
        ).resolve()
        if temporary.parent != parent:
            _reject(StandardExportErrorCode.IO_ERROR, "temporary")
        for path, content in sorted(package.files.items()):
            if path != "manifest.json":
                file_writer(temporary / path, content)
        file_writer(temporary / "manifest.json", package.files["manifest.json"])
        os.replace(temporary, destination)
        return destination
    except StandardExportError:
        raise
    except Exception as error:
        if destination.exists() and _directory_matches(destination, package):
            return destination
        raise StandardExportError(StandardExportErrorCode.IO_ERROR, field="write") from error
    finally:
        if temporary is not None and temporary.exists() and temporary.parent == parent:
            shutil.rmtree(temporary, ignore_errors=True)


__all__ = [
    "CHANGE_REPORT_READ_PROFILE",
    "CSV_DIALECT_VERSION",
    "EXPORT_JOB_VERSION",
    "EXPORT_MANIFEST_VERSION",
    "EXPORT_PACKAGE_PROFILE",
    "EXPORT_SCHEMA_SET_VERSION",
    "XLSX_PROFILE_VERSION",
    "ZIP_PROFILE_VERSION",
    "ChangeReportExportPackage",
    "archive_change_report_export_package",
    "build_change_report_export_package",
    "change_report_export_bytes_fingerprint",
    "load_change_report_export_package",
    "verify_change_report_export_package",
    "write_change_report_export_package",
]
