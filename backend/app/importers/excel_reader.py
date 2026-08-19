"""Read-only XLSX transport reader with active-content and archive guards."""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from io import BytesIO
from pathlib import PurePosixPath
import re
from typing import Any
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.importers.adapter import (
    AdapterErrorCode,
    InputAdapterError,
    REFERENCE_SHEET_NAME,
    ReferenceFileLimits,
    ReferenceTable,
    SourceCellRow,
    validate_reference_table,
)

_EXTERNAL_RELATIONSHIP = re.compile(
    rb"targetmode\s*=\s*['\"]external['\"]",
    flags=re.IGNORECASE,
)
_MACRO_CONTENT_MARKERS = (
    b"application/vnd.ms-office.vbaproject",
    b"macroenabled",
)


def _xlsx_error(
    code: AdapterErrorCode,
    *,
    location: str,
    expected: str,
    message: str,
) -> InputAdapterError:
    return InputAdapterError(
        code,
        source_location=location,
        expected_contract=expected,
        message=message,
    )


def _safe_archive_name(name: str) -> str:
    normalized = name.replace("\\", "/")
    path = PurePosixPath(normalized)
    if not normalized or normalized.startswith("/") or ".." in path.parts:
        raise _xlsx_error(
            AdapterErrorCode.UNSAFE_ARCHIVE,
            location="xlsx:archive",
            expected="relative OOXML member paths without traversal",
            message="the workbook archive contains an unsafe member path",
        )
    return normalized


def _preflight_archive(content: bytes, limits: ReferenceFileLimits) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > limits.max_archive_members:
                raise _xlsx_error(
                    AdapterErrorCode.UNSAFE_ARCHIVE,
                    location="xlsx:archive",
                    expected=f"at most {limits.max_archive_members} archive members",
                    message="the workbook archive has too many members",
                )

            total_size = 0
            names: set[str] = set()
            for member in members:
                normalized = _safe_archive_name(member.filename)
                identity = normalized.casefold()
                if identity in names:
                    raise _xlsx_error(
                        AdapterErrorCode.UNSAFE_ARCHIVE,
                        location="xlsx:archive",
                        expected="unique OOXML member paths",
                        message="the workbook archive contains duplicate member paths",
                    )
                names.add(identity)
                if member.flag_bits & 0x1:
                    raise _xlsx_error(
                        AdapterErrorCode.UNSAFE_ARCHIVE,
                        location="xlsx:archive",
                        expected="an unencrypted OOXML workbook",
                        message="encrypted workbook members are forbidden",
                    )
                total_size += member.file_size
                if total_size > limits.max_archive_uncompressed_bytes:
                    raise _xlsx_error(
                        AdapterErrorCode.UNSAFE_ARCHIVE,
                        location="xlsx:archive",
                        expected=(
                            "at most "
                            f"{limits.max_archive_uncompressed_bytes} uncompressed bytes"
                        ),
                        message="the workbook archive exceeds the expansion limit",
                    )

                lower_name = normalized.casefold()
                if lower_name.endswith("vbaproject.bin"):
                    raise _xlsx_error(
                        AdapterErrorCode.FORBIDDEN_MACRO,
                        location="xlsx:archive",
                        expected="an XLSX package without VBA content",
                        message="macro content is forbidden",
                    )
                if lower_name.startswith("xl/externallinks/"):
                    raise _xlsx_error(
                        AdapterErrorCode.FORBIDDEN_EXTERNAL_LINK,
                        location="xlsx:archive",
                        expected="an XLSX package without external links",
                        message="external workbook links are forbidden",
                    )

                if lower_name.endswith((".xml", ".rels")):
                    payload = archive.read(member)
                    lower_payload = payload.lower()
                    if b"<!doctype" in lower_payload or b"<!entity" in lower_payload:
                        raise _xlsx_error(
                            AdapterErrorCode.UNSAFE_ARCHIVE,
                            location="xlsx:archive",
                            expected="OOXML without DTD or entity declarations",
                            message="unsafe XML declarations are forbidden",
                        )
                    if any(
                        marker in lower_payload for marker in _MACRO_CONTENT_MARKERS
                    ):
                        raise _xlsx_error(
                            AdapterErrorCode.FORBIDDEN_MACRO,
                            location="xlsx:archive",
                            expected="an XLSX package without macro content types",
                            message="macro-enabled workbook content is forbidden",
                        )
                    if _EXTERNAL_RELATIONSHIP.search(payload):
                        raise _xlsx_error(
                            AdapterErrorCode.FORBIDDEN_EXTERNAL_LINK,
                            location="xlsx:archive",
                            expected="OOXML relationships without external targets",
                            message="external relationships are forbidden",
                        )

            required_members = {"[content_types].xml", "xl/workbook.xml"}
            if not required_members.issubset(names):
                raise _xlsx_error(
                    AdapterErrorCode.INVALID_WORKBOOK,
                    location="xlsx:archive",
                    expected="a complete XLSX OOXML package",
                    message="required workbook members are missing",
                )
    except InputAdapterError:
        raise
    except (BadZipFile, KeyError, NotImplementedError, OSError, RuntimeError):
        raise _xlsx_error(
            AdapterErrorCode.INVALID_WORKBOOK,
            location="xlsx:archive",
            expected="a readable XLSX OOXML package",
            message="the source is not a valid workbook archive",
        ) from None


def _text_row(cells: Sequence[Any], *, location: str) -> SourceCellRow:
    values: list[str] = []
    for cell in cells:
        if getattr(cell, "data_type", None) == "f":
            raise _xlsx_error(
                AdapterErrorCode.FORBIDDEN_FORMULA,
                location=location,
                expected="literal text cells without formulas",
                message="workbook formulas are forbidden",
            )
        value = getattr(cell, "value", None)
        if value is None:
            values.append("")
        elif isinstance(value, str):
            values.append(value)
        else:
            raise _xlsx_error(
                AdapterErrorCode.INVALID_CELL_TYPE,
                location=location,
                expected="text cells in every reference column",
                message="the workbook contains a non-text reference cell",
            )
    return SourceCellRow(source_location=location, values=tuple(values))


def read_reference_xlsx(
    content: bytes,
    *,
    limits: ReferenceFileLimits,
) -> ReferenceTable:
    """Read a single exact-name sheet without formulas, links, or active content."""

    _preflight_archive(content, limits)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (
        DefusedXmlException,
        InvalidFileException,
        ParseError,
        KeyError,
        OSError,
        TypeError,
        ValueError,
    ):
        raise _xlsx_error(
            AdapterErrorCode.INVALID_WORKBOOK,
            location="xlsx:workbook",
            expected="a safely parseable XLSX workbook",
            message="the workbook structure is invalid",
        ) from None

    try:
        if len(workbook.sheetnames) > limits.max_sheets:
            raise _xlsx_error(
                AdapterErrorCode.SHEET_LIMIT_EXCEEDED,
                location="xlsx:workbook",
                expected=f"at most {limits.max_sheets} worksheet",
                message="the workbook exceeds the configured sheet limit",
            )
        if workbook.sheetnames != [REFERENCE_SHEET_NAME]:
            raise _xlsx_error(
                AdapterErrorCode.INVALID_SHEET,
                location="xlsx:workbook",
                expected=f"one worksheet named {REFERENCE_SHEET_NAME}",
                message="the workbook sheet contract does not match adapter v1",
            )
        worksheet = workbook[REFERENCE_SHEET_NAME]
        if worksheet.max_column > limits.max_columns:
            raise _xlsx_error(
                AdapterErrorCode.COLUMN_LIMIT_EXCEEDED,
                location=f"xlsx:sheet={REFERENCE_SHEET_NAME}",
                expected=f"at most {limits.max_columns} columns",
                message="the worksheet exceeds the configured column limit",
            )
        if max(0, worksheet.max_row - 1) > limits.max_rows:
            raise _xlsx_error(
                AdapterErrorCode.ROW_LIMIT_EXCEEDED,
                location=f"xlsx:sheet={REFERENCE_SHEET_NAME}",
                expected=f"at most {limits.max_rows} data rows",
                message="the worksheet exceeds the configured row limit",
            )

        source_rows = worksheet.iter_rows()
        first = next(source_rows, None)
        header = (
            _text_row(
                first,
                location=f"xlsx:sheet={REFERENCE_SHEET_NAME},row=1",
            )
            if first is not None and any(cell.value is not None for cell in first)
            else None
        )

        def rows() -> Iterator[SourceCellRow]:
            for row_number, cells in enumerate(source_rows, start=2):
                if row_number - 1 > limits.max_rows:
                    raise _xlsx_error(
                        AdapterErrorCode.ROW_LIMIT_EXCEEDED,
                        location=(
                            f"xlsx:sheet={REFERENCE_SHEET_NAME},row={row_number}"
                        ),
                        expected=f"at most {limits.max_rows} data rows",
                        message="the worksheet exceeds the configured row limit",
                    )
                yield _text_row(
                    cells,
                    location=f"xlsx:sheet={REFERENCE_SHEET_NAME},row={row_number}",
                )

        return validate_reference_table(header=header, rows=rows(), limits=limits)
    finally:
        workbook.close()


__all__ = ["read_reference_xlsx"]
