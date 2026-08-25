"""Deterministic P3 standard package with a safe XLSX and manifest-last writes."""

from __future__ import annotations

from collections.abc import Callable, Mapping
import csv
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import StrEnum
from hashlib import sha256
import io
import json
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any, Never, cast
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.workspace_contracts import (
    canonical_workspace_bytes,
    require_workspace_document,
)
from app.exporters.package import (
    InternalExportPackage,
    verify_internal_export_package,
)


EXPORT_MANIFEST_VERSION = "export-manifest.v2"
EXPORT_JOB_VERSION = "export-job.v2"
EXPORT_SCHEMA_SET_VERSION = "2.7.0"
EXPORT_PACKAGE_PROFILE = "p3-standard-export.v1"
XLSX_PROFILE_VERSION = "xlsx-safe-deterministic.v1"
CSV_DIALECT_VERSION = "rfc4180-lf.v1"

_PAYLOAD_ROLES = {
    "schedule_version.json": "PUBLISHED_SCHEDULE_VERSION",
    "planning_solution.json": "VALIDATED_PLANNING_SOLUTION",
    "schedule_operations.csv": "SCHEDULE_OPERATIONS",
    "order_summary.csv": "ORDER_SUMMARY",
    "resource_load.csv": "RESOURCE_LOAD",
    "kpi.json": "KPI",
    "validation_report.json": "VALIDATION_REPORT",
    "solver_report.json": "SOLVER_REPORT",
    "import_quality_report.json": "IMPORT_QUALITY_REPORT",
    "scenario_manifest.json": "P2_CORRECTNESS_SCENARIO_MANIFEST",
    "publication_result.json": "PUBLICATION_RESULT",
    "standard_package.xlsx": "STANDARD_WORKBOOK",
}
_P2_TO_P3_PATHS = {
    "schedule.json": "planning_solution.json",
    "schedule_operations.csv": "schedule_operations.csv",
    "order_summary.csv": "order_summary.csv",
    "resource_load.csv": "resource_load.csv",
    "kpi.json": "kpi.json",
    "validation_report.json": "validation_report.json",
    "solver_report.json": "solver_report.json",
    "import_quality_report.json": "import_quality_report.json",
    "scenario_manifest.json": "scenario_manifest.json",
}
_DEFERRED = [
    {"path": "benchmark_report.json", "status": "NOT_APPLICABLE_STANDARD_EXPORT"},
    {"path": "change_report.json", "status": "DEFERRED_P4_DYNAMIC_REPLAN"},
]
_WORKBOOK_SHEETS = (
    ("Schedule Operations", "schedule_operations.csv"),
    ("Order Summary", "order_summary.csv"),
    ("Resource Load", "resource_load.csv"),
)

type JsonObject = dict[str, Any]


class StandardExportErrorCode(StrEnum):
    INVALID_PACKAGE = "INVALID_PACKAGE"
    MIXED_LINEAGE = "MIXED_LINEAGE"
    HASH_MISMATCH = "HASH_MISMATCH"
    UNSAFE_XLSX = "UNSAFE_XLSX"
    DESTINATION_CONFLICT = "DESTINATION_CONFLICT"
    IO_ERROR = "IO_ERROR"


class StandardExportError(ValueError):
    def __init__(self, code: StandardExportErrorCode, *, field: str) -> None:
        self.code = code
        self.field = field
        super().__init__(f"{code.value}: {field}")


@dataclass(frozen=True, slots=True)
class StandardExportPackage:
    package_id: str
    manifest_fingerprint: str
    storage_reference: str
    _files: tuple[tuple[str, bytes], ...]

    @property
    def files(self) -> dict[str, bytes]:
        return dict(self._files)

    @property
    def manifest(self) -> JsonObject:
        return cast(JsonObject, json.loads(self.files["manifest.json"]))


def _reject(code: StandardExportErrorCode, field: str) -> Never:
    raise StandardExportError(code, field=field)


def _fingerprint(value: bytes) -> str:
    return f"sha256:{sha256(value).hexdigest()}"


def _json_bytes(value: Mapping[str, object]) -> bytes:
    return canonical_workspace_bytes(value)


def _mapping(value: object, field: str) -> Mapping[str, object]:
    if not isinstance(value, Mapping):
        _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
    return cast(Mapping[str, object], value)


def _csv_rows(value: bytes, field: str) -> list[list[str]]:
    try:
        text = value.decode("utf-8")
    except UnicodeDecodeError:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
    if "\r" in text:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
    rows = list(csv.reader(io.StringIO(text, newline="")))
    if not rows:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, field)
    return rows


def _safe_cell(value: str) -> str:
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _canonicalize_xlsx(value: bytes) -> bytes:
    source = io.BytesIO(value)
    target = io.BytesIO()
    with ZipFile(source, "r") as archive, ZipFile(
        target, "w", compression=ZIP_DEFLATED, compresslevel=9
    ) as result:
        for name in sorted(archive.namelist()):
            content = archive.read(name)
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            result.writestr(info, content)
    return target.getvalue()


def _workbook_bytes(payloads: Mapping[str, bytes], metadata: Mapping[str, object]) -> bytes:
    workbook = Workbook()
    workbook.properties.creator = "PlantNexus APS"
    workbook.properties.lastModifiedBy = "PlantNexus APS"
    workbook.properties.created = datetime(1980, 1, 1, tzinfo=UTC)
    workbook.properties.modified = datetime(1980, 1, 1, tzinfo=UTC)
    workbook.properties.title = "P3 Standard Export"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    first = True
    for title, path in _WORKBOOK_SHEETS:
        sheet = cast(Worksheet, workbook.active) if first else workbook.create_sheet()
        first = False
        sheet.title = title
        for row in _csv_rows(payloads[path], path):
            sheet.append([_safe_cell(value) for value in row])
    sheet = workbook.create_sheet("Metadata")
    sheet.append(["key", "value"])
    for key, value in sorted(metadata.items()):
        rendered = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        sheet.append([_safe_cell(str(key)), _safe_cell(rendered)])
    stream = io.BytesIO()
    workbook.save(stream)
    rendered = _canonicalize_xlsx(stream.getvalue())
    _verify_xlsx(rendered)
    return rendered


def _verify_xlsx(value: bytes) -> int:
    try:
        with ZipFile(io.BytesIO(value), "r") as archive:
            names = set(archive.namelist())
            forbidden = (
                "vbaProject.bin",
                "externalLinks/",
                "embeddings/",
                "connections.xml",
            )
            if any(any(token in name for token in forbidden) for name in names):
                _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
            for name in names:
                if name.endswith(".xml") and re.search(br"<f(?:[ >])", archive.read(name)):
                    _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
        workbook = load_workbook(io.BytesIO(value), read_only=False, data_only=False)
        if tuple(workbook.sheetnames) != tuple(
            [title for title, _ in _WORKBOOK_SHEETS] + ["Metadata"]
        ):
            _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")
        return len(workbook.sheetnames)
    except StandardExportError:
        raise
    except Exception:
        _reject(StandardExportErrorCode.UNSAFE_XLSX, "standard_package.xlsx")


def _file_record(path: str, content: bytes) -> JsonObject:
    is_csv = path.endswith(".csv")
    is_xlsx = path.endswith(".xlsx")
    return {
        "path": path,
        "role": _PAYLOAD_ROLES[path],
        "media_type": (
            "text/csv; charset=utf-8"
            if is_csv
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if is_xlsx
            else "application/json"
        ),
        "sha256": _fingerprint(content),
        "size_bytes": len(content),
        "row_count": len(_csv_rows(content, path)) - 1 if is_csv else None,
        "sheet_count": _verify_xlsx(content) if is_xlsx else None,
    }


def storage_reference_for(package_id: str) -> str:
    return _fingerprint(f"SIMULATION_INTERNAL/{package_id}".encode("utf-8"))


def build_standard_export_package(
    *,
    p2_package: InternalExportPackage,
    schedule_version: Mapping[str, object],
    publication_result: Mapping[str, object],
    export_job: Mapping[str, object],
    create_audit_event_id: str,
    attempt_audit_event_id: str,
    completion_audit_event_id: str,
    correlation_id: str,
    generated_at_utc: str,
) -> StandardExportPackage:
    """Upgrade a frozen P2 package without rewriting any P2 payload bytes."""

    verify_internal_export_package(p2_package)
    if require_workspace_document(schedule_version) != "schedule-version.v1":
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "schedule_version")
    if schedule_version.get("state") != "PUBLISHED":
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "schedule_version.state")
    if require_workspace_document(publication_result) != "publication-result.v1":
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "publication_result")
    if require_workspace_document(export_job) != EXPORT_JOB_VERSION:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "export_job")
    if export_job.get("state") != "EXPORTING" or not isinstance(export_job.get("attempt"), int) or cast(int, export_job["attempt"]) < 1:
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "export_job.state/attempt")

    schedule_reference = _mapping(export_job.get("schedule_version"), "export_job.schedule_version")
    published_reference = _mapping(publication_result.get("published_version"), "publication_result.published_version")
    for field in ("schedule_version_id", "content_fingerprint"):
        expected = schedule_version.get(field)
        if schedule_reference.get(field) != expected or published_reference.get(field) != expected:
            _reject(StandardExportErrorCode.MIXED_LINEAGE, field)
    if (
        publication_result.get("target") != "SIMULATION_INTERNAL"
        or export_job.get("target") != "SIMULATION_INTERNAL"
        or schedule_version.get("data_plane") != "SIMULATION"
        or publication_result.get("data_plane") != "SIMULATION"
        or export_job.get("data_plane") != "SIMULATION"
        or schedule_version.get("synthetic") is not True
    ):
        _reject(StandardExportErrorCode.MIXED_LINEAGE, "plane/target/synthetic")

    p2_files = p2_package.files
    p2_manifest = p2_package.manifest
    try:
        planning_solution = cast(JsonObject, json.loads(p2_files["schedule.json"]))
        content = _mapping(schedule_version.get("content"), "schedule_version.content")
        lineage = _mapping(schedule_version.get("lineage"), "schedule_version.lineage")
        schedule_solution = _mapping(lineage.get("planning_solution"), "schedule_version.lineage.planning_solution")
        p2_solution = _mapping(_mapping(p2_manifest.get("lineage"), "p2.lineage").get("solution"), "p2.lineage.solution")
        if content.get("assignments") != planning_solution.get("assignments") or schedule_solution.get("fingerprint") != p2_solution.get("solution_fingerprint"):
            _reject(StandardExportErrorCode.MIXED_LINEAGE, "schedule_version.lineage/content")
    except (UnicodeDecodeError, json.JSONDecodeError):
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "planning_solution.json")
    payloads = {
        target: p2_files[source] for source, target in _P2_TO_P3_PATHS.items()
    }
    payloads["schedule_version.json"] = _json_bytes(schedule_version)
    payloads["publication_result.json"] = _json_bytes(publication_result)
    payloads["standard_package.xlsx"] = _workbook_bytes(
        payloads,
        {
            "export_job_id": export_job["export_job_id"],
            "attempt": export_job["attempt"],
            "schedule_version_id": schedule_version["schedule_version_id"],
            "content_fingerprint": schedule_version["content_fingerprint"],
            "publication_id": publication_result["publication_id"],
            "target": "SIMULATION_INTERNAL",
            "profile": EXPORT_PACKAGE_PROFILE,
        },
    )
    file_records = [_file_record(path, payloads[path]) for path in sorted(payloads)]
    manifest_basis: JsonObject = {
        "export_manifest_version": EXPORT_MANIFEST_VERSION,
        "schema_set_version": EXPORT_SCHEMA_SET_VERSION,
        "package_profile": EXPORT_PACKAGE_PROFILE,
        "canonicalization_version": "canonical-json.v1",
        "csv_dialect_version": CSV_DIALECT_VERSION,
        "xlsx_profile_version": XLSX_PROFILE_VERSION,
        "generated_at_utc": generated_at_utc,
        "publishable": False,
        "target": "SIMULATION_INTERNAL",
        "schedule_version": {
            "schedule_version_id": schedule_version["schedule_version_id"],
            "state": "PUBLISHED",
            "content_fingerprint": schedule_version["content_fingerprint"],
        },
        "publication": {
            "publication_result_version": publication_result["publication_result_version"],
            "publication_id": publication_result["publication_id"],
            "result_fingerprint": publication_result["result_fingerprint"],
            "published_at_utc": publication_result["published_at_utc"],
            "audit_event_id": publication_result["audit_event_id"],
        },
        "export_job": {
            "export_job_version": EXPORT_JOB_VERSION,
            "export_job_id": export_job["export_job_id"],
            "attempt": export_job["attempt"],
            "state_at_materialization": "EXPORTING",
        },
        "audit_lineage": {
            "create_audit_event_id": create_audit_event_id,
            "attempt_audit_event_id": attempt_audit_event_id,
            "completion_audit_event_id": completion_audit_event_id,
            "correlation_id": correlation_id,
        },
        "p2_package": {
            "export_manifest_version": p2_manifest["export_manifest_version"],
            "package_profile": p2_manifest["package_profile"],
            "package_id": p2_package.package_id,
            "manifest_fingerprint": p2_package.manifest_fingerprint,
        },
        "p2_lineage": p2_manifest["lineage"],
        "entity_counts": p2_manifest["entity_counts"],
        "file_count": len(file_records),
        "files": file_records,
        "deferred_artifacts": _DEFERRED,
        "state_boundary": {
            "schedule_version": "PUBLISHED",
            "publication": "COMPLETED",
            "export_job_at_materialization": "EXPORTING",
            "external_transfer": "NOT_STARTED",
            "production": "NOT_AUTHORIZED",
        },
        "synthetic": True,
        "synthetic_provenance": schedule_version["synthetic_provenance"],
    }
    package_id = f"export-package-{sha256(_json_bytes(manifest_basis)).hexdigest()}"
    manifest = {"package_id": package_id, **manifest_basis}
    manifest_bytes = _json_bytes(manifest)
    package = StandardExportPackage(
        package_id=package_id,
        manifest_fingerprint=_fingerprint(manifest_bytes),
        storage_reference=storage_reference_for(package_id),
        _files=tuple(sorted({"manifest.json": manifest_bytes, **payloads}.items())),
    )
    verify_standard_export_package(package)
    return package


def verify_standard_export_package(package: StandardExportPackage) -> None:
    try:
        files = package.files
        if set(files) != {"manifest.json", *_PAYLOAD_ROLES}:
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "files")
        manifest = cast(JsonObject, json.loads(files["manifest.json"]))
        if (
            manifest.get("export_manifest_version") != EXPORT_MANIFEST_VERSION
            or manifest.get("schema_set_version") != EXPORT_SCHEMA_SET_VERSION
            or manifest.get("package_profile") != EXPORT_PACKAGE_PROFILE
            or manifest.get("file_count") != len(_PAYLOAD_ROLES)
            or manifest.get("deferred_artifacts") != _DEFERRED
        ):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "manifest")
        basis = {key: value for key, value in manifest.items() if key != "package_id"}
        expected_id = f"export-package-{sha256(_json_bytes(basis)).hexdigest()}"
        if package.package_id != expected_id or manifest.get("package_id") != expected_id:
            _reject(StandardExportErrorCode.HASH_MISMATCH, "package_id")
        if package.manifest_fingerprint != _fingerprint(files["manifest.json"]):
            _reject(StandardExportErrorCode.HASH_MISMATCH, "manifest_fingerprint")
        if package.storage_reference != storage_reference_for(package.package_id):
            _reject(StandardExportErrorCode.HASH_MISMATCH, "storage_reference")
        records = cast(list[Mapping[str, object]], manifest["files"])
        if [record.get("path") for record in records] != sorted(_PAYLOAD_ROLES):
            _reject(StandardExportErrorCode.INVALID_PACKAGE, "manifest.files")
        for record in records:
            path = cast(str, record["path"])
            content = files[path]
            if record.get("role") != _PAYLOAD_ROLES[path] or record.get("sha256") != _fingerprint(content) or record.get("size_bytes") != len(content):
                _reject(StandardExportErrorCode.HASH_MISMATCH, path)
            if path.endswith(".csv") and record.get("row_count") != len(_csv_rows(content, path)) - 1:
                _reject(StandardExportErrorCode.HASH_MISMATCH, path)
            if path.endswith(".xlsx") and record.get("sheet_count") != _verify_xlsx(content):
                _reject(StandardExportErrorCode.UNSAFE_XLSX, path)
    except StandardExportError:
        raise
    except Exception:
        _reject(StandardExportErrorCode.INVALID_PACKAGE, "package")


def _directory_matches(destination: Path, package: StandardExportPackage) -> bool:
    if not destination.is_dir():
        return False
    children = list(destination.iterdir())
    return all(path.is_file() for path in children) and {
        path.name: path.read_bytes() for path in children
    } == package.files


def _write_file(path: Path, value: bytes) -> None:
    path.write_bytes(value)


def write_standard_export_package(
    package: StandardExportPackage,
    destination: Path,
    *,
    file_writer: Callable[[Path, bytes], None] = _write_file,
) -> Path:
    """Write payloads first and manifest last, then atomically rename the directory."""

    verify_standard_export_package(package)
    destination = destination.resolve()
    parent = destination.parent
    temporary: Path | None = None
    try:
        if not parent.is_dir():
            _reject(StandardExportErrorCode.IO_ERROR, "destination.parent")
        if destination.exists():
            if _directory_matches(destination, package):
                return destination
            _reject(StandardExportErrorCode.DESTINATION_CONFLICT, "destination")
        temporary = Path(tempfile.mkdtemp(prefix=f".{destination.name}.tmp-", dir=parent)).resolve()
        if temporary.parent != parent:
            _reject(StandardExportErrorCode.IO_ERROR, "temporary")
        for path, content in sorted(package.files.items()):
            if path != "manifest.json":
                file_writer(temporary / path, content)
        file_writer(temporary / "manifest.json", package.files["manifest.json"])
        os.replace(temporary, destination)
        return destination
    except StandardExportError:
        raise
    except Exception as error:
        if destination.exists() and _directory_matches(destination, package):
            return destination
        raise StandardExportError(StandardExportErrorCode.IO_ERROR, field="write") from error
    finally:
        if temporary is not None and temporary.exists() and temporary.parent == parent:
            shutil.rmtree(temporary, ignore_errors=True)


__all__ = [
    "EXPORT_JOB_VERSION", "EXPORT_MANIFEST_VERSION", "EXPORT_PACKAGE_PROFILE",
    "EXPORT_SCHEMA_SET_VERSION", "StandardExportError", "StandardExportErrorCode",
    "StandardExportPackage", "build_standard_export_package", "storage_reference_for",
    "verify_standard_export_package", "write_standard_export_package",
]
