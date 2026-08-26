"""TEST-OUTPUT: deterministic P3 JSON/CSV/XLSX package and atomic storage."""

from __future__ import annotations

from collections.abc import Mapping
from copy import deepcopy
from datetime import UTC, datetime
import io
import json
from pathlib import Path
from typing import Any, cast
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest

from app.domain.export_job import (
    ExportJobContext,
    ExportJobError,
    ExportJobRequest,
    audit_event_id,
    build_created_export_job,
    export_job_identity,
    lease_reference_for,
    transition_export_job,
)
from app.application.export_jobs import ExportJobServiceResult
from app.application.export_downloads import ExportPackageDownloadService
from app.domain.workspace_contracts import (
    publication_result_fingerprint,
    require_workspace_document,
    schedule_content_fingerprint,
)
from app.exporters import InternalExportPackage, build_internal_export_package
from app.exporters.standard_package import (
    StandardExportError,
    StandardExportErrorCode,
    StandardExportPackage,
    archive_standard_export_package,
    build_standard_export_package,
    load_standard_export_package,
    standard_export_bytes_fingerprint,
    verify_standard_export_package,
    write_standard_export_package,
)
from app.simulation.scenarios.p2_correctness import (
    execute_correctness_case,
    load_correctness_cases,
)
from app.jobs.export_job import InternalExportJobWorker
from app.jobs.export_package_store import (
    LocalExportPackageStore,
    StoredVerifiedExportPackage,
    export_attempt_destination,
)


ROOT = Path(__file__).resolve().parents[3]


@pytest.fixture(scope="module")
def package_inputs() -> tuple[
    InternalExportPackage, dict[str, object], dict[str, object], dict[str, object]
]:
    replay = execute_correctness_case(load_correctness_cases(ROOT)[0], root=ROOT)
    p2 = build_internal_export_package(
        snapshot=replay.snapshot_document,
        problem=replay.problem,
        solution=replay.solution,
        solver_report=replay.solver_report,
        validation_report=replay.validation_report,
        import_quality_report=replay.quality_report,
        scenario_manifest=replay.case.manifest,
    )
    schedule = json.loads(
        (ROOT / "schemas/samples/schedule-version.v1.synthetic.json").read_text(
            encoding="utf-8"
        )
    )
    schedule["state"] = "PUBLISHED"
    schedule["content"] = {"assignments": replay.solution["assignments"], "locks": []}
    schedule["content_fingerprint"] = schedule_content_fingerprint(schedule)
    schedule["lineage"]["planning_solution"] = {
        "document_version": "planning-solution.v1",
        "artifact_id": replay.solution["solution_id"],
        "fingerprint": p2.manifest["lineage"]["solution"]["solution_fingerprint"],
    }
    schedule["decision"] = {
        "decision": "APPROVED",
        "actor_ref": "actor:sim-approver-001",
        "capability": "approve",
        "reason": "Synthetic export fixture approval.",
        "decided_at_utc": "2026-08-25T00:00:00Z",
        "audit_event_id": "audit-approve-export-fixture",
    }
    schedule["publication"] = {
        "publication_id": "publication-export-fixture",
        "target": "SIMULATION_INTERNAL",
        "published_at_utc": "2026-08-25T00:01:00Z",
        "audit_event_id": "audit-publish-export-fixture",
    }
    schedule["allowed_actions"] = ["view", "export"]
    require_workspace_document(schedule)

    publication = json.loads(
        (ROOT / "schemas/samples/publication-result.v1.synthetic.json").read_text(
            encoding="utf-8"
        )
    )
    reference = {
        "schedule_version_id": schedule["schedule_version_id"],
        "state": "PUBLISHED",
        "content_fingerprint": schedule["content_fingerprint"],
    }
    publication["publication_id"] = "publication-export-fixture"
    publication["source_approved_version"] = {**reference, "state": "APPROVED"}
    publication["published_version"] = reference
    publication["published_at_utc"] = "2026-08-25T00:01:00Z"
    publication["audit_event_id"] = "audit-publish-export-fixture"
    publication["synthetic_provenance"] = deepcopy(schedule["synthetic_provenance"])
    publication["result_fingerprint"] = publication_result_fingerprint(publication)
    require_workspace_document(publication)

    request = ExportJobRequest(
        schedule_version_id=str(schedule["schedule_version_id"]),
        expected_content_fingerprint=str(schedule["content_fingerprint"]),
        raw_idempotency_key="export-fixture-key-0001",
        reason="Synthetic standard package test.",
        correlation_id="correlation-export-fixture",
        environment="TEST",
        synthetic_provenance=deepcopy(schedule["synthetic_provenance"]),
    )
    context = ExportJobContext(
        actor_ref="actor:export-worker",
        authenticated=True,
        resolved_capabilities=frozenset({"export"}),
        schedule_version_scope=frozenset({str(schedule["schedule_version_id"])}),
        export_job_scope=frozenset(),
        auth_policy_version="sim-policy.v1",
        production_binding=False,
        occurred_at_utc="2026-08-25T00:02:00Z",
        code_commit="uncommitted",
    )
    identity = export_job_identity(request)
    created = build_created_export_job(
        request, identity, context, schedule, publication
    )
    attempt = 1
    exporting = transition_export_job(
        created,
        target_state="EXPORTING",
        occurred_at_utc="2026-08-25T00:03:00Z",
        audit_event_id_value=audit_event_id(identity.export_job_id, "ATTEMPT", attempt),
        attempt=attempt,
        lease_reference=lease_reference_for(
            identity.export_job_id, attempt, "worker:test"
        ),
    )
    return p2, schedule, publication, exporting


def _build(
    inputs: tuple[
        InternalExportPackage, dict[str, object], dict[str, object], dict[str, object]
    ],
):  # type: ignore[no-untyped-def]
    p2, schedule, publication, job = inputs
    attempt = int(str(job["attempt"]))
    return build_standard_export_package(
        p2_package=p2,
        schedule_version=schedule,
        publication_result=publication,
        export_job=job,
        create_audit_event_id=audit_event_id(str(job["export_job_id"]), "CREATE", 0),
        attempt_audit_event_id=str(job["latest_audit_event_id"]),
        completion_audit_event_id=audit_event_id(
            str(job["export_job_id"]), "COMPLETED", attempt
        ),
        correlation_id="correlation-export-fixture",
        generated_at_utc="2026-08-25T00:04:00Z",
    )


def test_standard_package_is_byte_deterministic_and_preserves_p2_payloads(
    package_inputs,
) -> None:  # type: ignore[no-untyped-def]
    first = _build(package_inputs)
    second = _build(package_inputs)
    assert first == second
    verify_standard_export_package(first)
    p2 = package_inputs[0]
    assert first.files["planning_solution.json"] == p2.files["schedule.json"]
    for path in (
        "schedule_operations.csv",
        "order_summary.csv",
        "resource_load.csv",
        "kpi.json",
        "validation_report.json",
        "solver_report.json",
        "import_quality_report.json",
        "scenario_manifest.json",
    ):
        assert first.files[path] == p2.files[path]
    assert first.manifest["file_count"] == 12
    assert (
        first.manifest["deferred_artifacts"][1]["status"]
        == "DEFERRED_P4_DYNAMIC_REPLAN"
    )


def test_xlsx_has_fixed_safe_sheets_and_no_active_content(package_inputs) -> None:  # type: ignore[no-untyped-def]
    package = _build(package_inputs)
    value = package.files["standard_package.xlsx"]
    workbook = load_workbook(io.BytesIO(value), data_only=False)
    assert workbook.sheetnames == [
        "Schedule Operations",
        "Order Summary",
        "Resource Load",
        "Metadata",
    ]
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    with ZipFile(io.BytesIO(value)) as archive:
        assert all(
            "externalLinks/" not in name and "vbaProject.bin" not in name
            for name in archive.namelist()
        )


def test_tamper_and_mixed_lineage_fail_closed(package_inputs) -> None:  # type: ignore[no-untyped-def]
    package = _build(package_inputs)
    files = package.files
    files["kpi.json"] += b" "
    tampered = StandardExportPackage(
        package.package_id,
        package.manifest_fingerprint,
        package.storage_reference,
        tuple(sorted(files.items())),
    )
    with pytest.raises(StandardExportError) as captured:
        verify_standard_export_package(tampered)
    assert captured.value.code is StandardExportErrorCode.HASH_MISMATCH

    _, schedule, publication, _ = package_inputs
    wrong_schedule = deepcopy(schedule)
    wrong_schedule["content"]["assignments"] = []
    wrong_schedule["content_fingerprint"] = schedule_content_fingerprint(wrong_schedule)
    with pytest.raises(StandardExportError) as mixed:
        build_standard_export_package(
            p2_package=package_inputs[0],
            schedule_version=wrong_schedule,
            publication_result=publication,
            export_job=package_inputs[3],
            create_audit_event_id="audit-create",
            attempt_audit_event_id="audit-attempt",
            completion_audit_event_id="audit-complete",
            correlation_id="correlation",
            generated_at_utc="2026-08-25T00:04:00Z",
        )
    assert mixed.value.code is StandardExportErrorCode.MIXED_LINEAGE


def test_manifest_last_atomic_replay_conflict_and_cleanup(
    package_inputs, tmp_path: Path
) -> None:  # type: ignore[no-untyped-def]
    package = _build(package_inputs)
    destination = tmp_path / "export-attempt"
    write_order: list[str] = []

    def recording_writer(path: Path, value: bytes) -> None:
        write_order.append(path.name)
        path.write_bytes(value)

    assert (
        write_standard_export_package(
            package, destination, file_writer=recording_writer
        )
        == destination
    )
    assert write_order[-1] == "manifest.json"
    assert write_standard_export_package(package, destination) == destination
    (destination / "kpi.json").write_bytes(b"tampered")
    with pytest.raises(StandardExportError) as conflict:
        write_standard_export_package(package, destination)
    assert conflict.value.code is StandardExportErrorCode.DESTINATION_CONFLICT

    failed_destination = tmp_path / "failed-attempt"

    def failing_writer(path: Path, value: bytes) -> None:
        if path.name == "resource_load.csv":
            raise OSError("injected")
        path.write_bytes(value)

    with pytest.raises(StandardExportError):
        write_standard_export_package(
            package, failed_destination, file_writer=failing_writer
        )
    assert not failed_destination.exists()
    assert not list(tmp_path.glob(".failed-attempt.tmp-*"))


def test_worker_composes_claim_package_storage_and_completion_without_publish(
    package_inputs, tmp_path: Path
) -> None:  # type: ignore[no-untyped-def]
    _, schedule, publication, exporting = package_inputs
    event_id = cast(str, exporting["latest_audit_event_id"])

    class FakeService:
        failed = False
        artifact: Mapping[str, object] | None = None

        def claim(self, *_args: object, **_kwargs: object) -> ExportJobServiceResult:
            return ExportJobServiceResult(exporting, 1, event_id, False)

        def complete(
            self,
            _job_id: str,
            context: ExportJobContext,
            *,
            expected_lease_reference: str,
            artifact_manifest: Mapping[str, object],
        ) -> ExportJobServiceResult:
            assert expected_lease_reference == exporting["lease_reference"]
            self.artifact = artifact_manifest
            completed = transition_export_job(
                exporting,
                target_state="EXPORTED",
                occurred_at_utc=context.occurred_at_utc,
                audit_event_id_value=audit_event_id(
                    str(exporting["export_job_id"]), "COMPLETED", 1
                ),
                artifact_manifest=artifact_manifest,
            )
            return ExportJobServiceResult(
                completed, 2, str(completed["latest_audit_event_id"]), False
            )

        def fail(self, *_args: object, **_kwargs: object) -> ExportJobServiceResult:
            self.failed = True
            raise AssertionError("worker success must not fail the job")

    fake = FakeService()
    context = ExportJobContext(
        actor_ref="actor:export-worker",
        authenticated=True,
        resolved_capabilities=frozenset({"export"}),
        schedule_version_scope=frozenset({str(schedule["schedule_version_id"])}),
        export_job_scope=frozenset({str(exporting["export_job_id"])}),
        auth_policy_version="sim-policy.v1",
        production_binding=False,
        occurred_at_utc="2026-08-25T00:04:00Z",
        code_commit="uncommitted",
    )
    worker = InternalExportJobWorker(service=cast(Any, fake), storage_root=tmp_path)
    result = worker.run(
        export_job_id=str(exporting["export_job_id"]),
        claim_context=context,
        terminal_context=context,
        owner_reference="worker:test",
        lease_expires_at_utc=datetime(2026, 8, 25, 0, 10, tzinfo=UTC),
        p2_package=package_inputs[0],
        schedule_version=schedule,
        publication_result=publication,
        correlation_id="correlation-export-fixture",
    )
    assert result.job.document["state"] == "EXPORTED"
    assert (
        result.destination.is_dir() and (result.destination / "manifest.json").is_file()
    )
    assert (
        fake.artifact is not None
        and fake.artifact["export_manifest_version"] == "export-manifest.v2"
    )
    assert fake.failed is False


def test_verified_loader_and_deterministic_archive_fail_closed(
    package_inputs, tmp_path: Path
) -> None:  # type: ignore[no-untyped-def]
    package = _build(package_inputs)
    destination = tmp_path / "verified-package"
    write_standard_export_package(package, destination)

    loaded = load_standard_export_package(destination)
    first = archive_standard_export_package(loaded)
    second = archive_standard_export_package(loaded)
    assert loaded == package
    assert first == second
    assert standard_export_bytes_fingerprint(first).startswith("sha256:")
    with ZipFile(io.BytesIO(first)) as archive:
        assert archive.namelist()[-1] == "manifest.json"
        assert {
            name: archive.read(name) for name in archive.namelist()
        } == package.files

    (destination / "unexpected.txt").write_text(
        "not part of the package", encoding="utf-8"
    )
    with pytest.raises(StandardExportError) as extra:
        load_standard_export_package(destination)
    assert extra.value.code is StandardExportErrorCode.INVALID_PACKAGE


def test_download_service_requires_authority_exposed_state_and_exact_artifact_lineage(
    package_inputs, tmp_path: Path
) -> None:  # type: ignore[no-untyped-def]
    package = _build(package_inputs)
    _, schedule, _, exporting = package_inputs
    export_job_id = cast(str, exporting["export_job_id"])
    attempt = cast(int, exporting["attempt"])
    destination = export_attempt_destination(
        tmp_path,
        export_job_id=export_job_id,
        attempt=attempt,
    )
    write_standard_export_package(package, destination)
    completed = transition_export_job(
        exporting,
        target_state="EXPORTED",
        occurred_at_utc="2026-08-25T00:05:00Z",
        audit_event_id_value=audit_event_id(export_job_id, "COMPLETED", attempt),
        artifact_manifest={
            "export_manifest_version": "export-manifest.v2",
            "package_id": package.package_id,
            "manifest_fingerprint": package.manifest_fingerprint,
            "storage_reference": package.storage_reference,
        },
    )

    class Stored:
        def __init__(self, document: dict[str, object]) -> None:
            self.document = document

    class Repository:
        def __init__(self, document: dict[str, object]) -> None:
            self.document = document
            self.calls = 0

        def get(self, export_job_id: str) -> Stored | None:
            self.calls += 1
            return Stored(self.document) if export_job_id == job_id else None

    job_id = export_job_id
    repository = Repository(completed)
    service = ExportPackageDownloadService(
        export_job_repository=repository,
        package_store=LocalExportPackageStore(tmp_path),
    )
    context = ExportJobContext(
        actor_ref="actor:export-downloader",
        authenticated=True,
        resolved_capabilities=frozenset({"export"}),
        schedule_version_scope=frozenset({cast(str, schedule["schedule_version_id"])}),
        export_job_scope=frozenset({export_job_id}),
        auth_policy_version="simulation-download-policy.v1",
        production_binding=False,
        occurred_at_utc="2026-08-25T00:06:00Z",
        code_commit="uncommitted",
    )
    result = service.download(
        export_job_id,
        context,
        correlation_id="correlation-export-download",
    )
    assert result.package_id == package.package_id
    assert result.manifest_fingerprint == package.manifest_fingerprint
    assert result.archive_fingerprint == standard_export_bytes_fingerprint(
        result.content
    )
    assert result.filename == f"{package.package_id}.zip"
    with ZipFile(io.BytesIO(result.content)) as archive:
        assert (
            json.loads(archive.read("manifest.json"))["package_id"]
            == package.package_id
        )

    class ForgedArchiveFingerprintStore:
        def load(
            self, *, export_job_id: str, attempt: int
        ) -> StoredVerifiedExportPackage:
            verified = cast(
                StoredVerifiedExportPackage,
                LocalExportPackageStore(tmp_path).load(
                    export_job_id=export_job_id,
                    attempt=attempt,
                ),
            )
            return StoredVerifiedExportPackage(
                content=verified.content,
                package_id=verified.package_id,
                manifest_fingerprint=verified.manifest_fingerprint,
                storage_reference=verified.storage_reference,
                archive_fingerprint="sha256:" + "0" * 64,
                manifest=verified.manifest,
            )

    with pytest.raises(ExportJobError) as forged_archive:
        ExportPackageDownloadService(
            export_job_repository=Repository(completed),
            package_store=ForgedArchiveFingerprintStore(),
        ).download(
            export_job_id,
            context,
            correlation_id="correlation-export-forged-archive",
        )
    assert forged_archive.value.reason.value == "EXPORT_FAILED"

    unauthorized_repository = Repository(completed)
    unauthorized = ExportPackageDownloadService(
        export_job_repository=unauthorized_repository,
        package_store=LocalExportPackageStore(tmp_path),
    )
    denied_context = ExportJobContext(
        actor_ref=context.actor_ref,
        authenticated=context.authenticated,
        resolved_capabilities=context.resolved_capabilities,
        schedule_version_scope=context.schedule_version_scope,
        export_job_scope=frozenset(),
        auth_policy_version=context.auth_policy_version,
        production_binding=context.production_binding,
        occurred_at_utc=context.occurred_at_utc,
        code_commit=context.code_commit,
    )
    with pytest.raises(ExportJobError) as denied:
        unauthorized.download(
            export_job_id,
            denied_context,
            correlation_id="correlation-export-denied",
        )
    assert denied.value.reason.value == "AUTHORIZATION_DENIED"
    assert unauthorized_repository.calls == 0

    mismatched = deepcopy(completed)
    cast(dict[str, object], mismatched["artifact_manifest"])["manifest_fingerprint"] = (
        "sha256:" + "0" * 64
    )
    with pytest.raises(ExportJobError) as mismatch:
        ExportPackageDownloadService(
            export_job_repository=Repository(mismatched),
            package_store=LocalExportPackageStore(tmp_path),
        ).download(
            export_job_id,
            context,
            correlation_id="correlation-export-mismatch",
        )
    assert mismatch.value.reason.value == "EXPORT_FAILED"

    wrong_completion_audit = deepcopy(completed)
    wrong_completion_audit["latest_audit_event_id"] = "audit-wrong-completion"
    with pytest.raises(ExportJobError) as audit_mismatch:
        ExportPackageDownloadService(
            export_job_repository=Repository(wrong_completion_audit),
            package_store=LocalExportPackageStore(tmp_path),
        ).download(
            export_job_id,
            context,
            correlation_id="correlation-export-audit-mismatch",
        )
    assert audit_mismatch.value.reason.value == "EXPORT_FAILED"
