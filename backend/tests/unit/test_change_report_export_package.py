"""TEST-OUTPUT: deterministic P4 ChangeReport package and atomic storage."""

from __future__ import annotations

from pathlib import Path
import io
from typing import Any, cast
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest

from app.exporters.change_report_output_check import (
    ChangeReportOutputFixture,
    build_change_report_output_fixture,
    build_fixture_package,
)
from app.exporters.change_report_package import (
    ChangeReportExportPackage,
    archive_change_report_export_package,
    load_change_report_export_package,
    verify_change_report_export_package,
    write_change_report_export_package,
)
from app.exporters.standard_package import StandardExportError, StandardExportErrorCode


ROOT = Path(__file__).resolve().parents[3]
REUSED_P3_PATHS = (
    "planning_solution.json",
    "schedule_operations.csv",
    "order_summary.csv",
    "resource_load.csv",
    "import_quality_report.json",
    "scenario_manifest.json",
)


@pytest.fixture(scope="module")
def fixture() -> ChangeReportOutputFixture:
    return build_change_report_output_fixture(ROOT)


@pytest.fixture(scope="module")
def package(fixture: ChangeReportOutputFixture) -> ChangeReportExportPackage:
    return build_fixture_package(fixture)


def test_package_is_byte_deterministic_manifest_bound_and_preserves_p3_payloads(
    fixture: ChangeReportOutputFixture,
    package: ChangeReportExportPackage,
) -> None:
    replay = build_fixture_package(fixture)
    assert replay == package
    verify_change_report_export_package(package)
    p3_files = cast(Any, fixture.p3_package).files
    for path in REUSED_P3_PATHS:
        assert package.files[path] == p3_files[path]

    manifest = package.manifest
    assert manifest["export_manifest_version"] == "export-manifest.v3"
    assert manifest["package_profile"] == "p4-dynamic-replan-export.v1"
    assert manifest["file_count"] == 13
    assert manifest["publishable"] is False
    assert manifest["target"] == "SIMULATION_INTERNAL"
    assert manifest["change_report"] == fixture.created_job["change_report"]
    assert manifest["p3_package"] == {
        "export_manifest_version": "export-manifest.v2",
        "package_profile": "p3-standard-export.v1",
        "package_id": cast(Any, fixture.p3_package).package_id,
        "manifest_fingerprint": cast(Any, fixture.p3_package).manifest_fingerprint,
    }


def test_workbook_has_five_fixed_safe_sheets_and_complete_change_rows(
    fixture: ChangeReportOutputFixture,
    package: ChangeReportExportPackage,
) -> None:
    value = package.files["standard_package.xlsx"]
    workbook = load_workbook(io.BytesIO(value), data_only=False)
    assert workbook.sheetnames == [
        "Schedule Operations",
        "Order Summary",
        "Resource Load",
        "Change Report",
        "Metadata",
    ]
    report_sheet = workbook["Change Report"]
    assert report_sheet.max_row == (
        cast(int, fixture.change_report["operation_universe_count"]) + 1
    )
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    with ZipFile(io.BytesIO(value)) as archive:
        assert all(
            "externalLinks/" not in name
            and "vbaProject.bin" not in name
            and "embeddings/" not in name
            for name in archive.namelist()
        )


def test_tamper_and_unmaterialized_job_fail_closed(
    fixture: ChangeReportOutputFixture,
    package: ChangeReportExportPackage,
) -> None:
    files = package.files
    files["change_report.json"] += b" "
    tampered = ChangeReportExportPackage(
        package.package_id,
        package.manifest_fingerprint,
        package.storage_reference,
        tuple(sorted(files.items())),
    )
    with pytest.raises(StandardExportError) as captured:
        verify_change_report_export_package(tampered)
    assert captured.value.code in {
        StandardExportErrorCode.HASH_MISMATCH,
        StandardExportErrorCode.INVALID_PACKAGE,
    }

    from app.exporters.change_report_package import build_change_report_export_package

    with pytest.raises(StandardExportError) as state_error:
        build_change_report_export_package(
            p3_package=cast(Any, fixture.p3_package),
            schedule_version=fixture.schedule_version,
            publication_result=fixture.publication_result,
            export_job=fixture.created_job,
            change_report=fixture.change_report,
            solver_report=fixture.solver_report,
            validation_report=fixture.validation_report,
            kpi=fixture.kpi,
            create_audit_event_id="audit-create",
            attempt_audit_event_id="audit-attempt",
            completion_audit_event_id="audit-complete",
            correlation_id=fixture.request.correlation_id,
            generated_at_utc="2026-08-28T10:11:00Z",
        )
    assert state_error.value.code is StandardExportErrorCode.MIXED_LINEAGE


def test_manifest_last_atomic_replay_conflict_cleanup_and_archive(
    package: ChangeReportExportPackage,
    tmp_path: Path,
) -> None:
    destination = tmp_path / "p4-export-attempt"
    order: list[str] = []

    def recording_writer(path: Path, value: bytes) -> None:
        order.append(path.name)
        path.write_bytes(value)

    assert (
        write_change_report_export_package(
            package,
            destination,
            file_writer=recording_writer,
        )
        == destination
    )
    assert order[-1] == "manifest.json"
    assert write_change_report_export_package(package, destination) == destination
    assert load_change_report_export_package(destination) == package
    assert archive_change_report_export_package(package) == (
        archive_change_report_export_package(package)
    )

    (destination / "change_report.json").write_bytes(b"tampered")
    with pytest.raises(StandardExportError) as conflict:
        write_change_report_export_package(package, destination)
    assert conflict.value.code is StandardExportErrorCode.DESTINATION_CONFLICT

    failed = tmp_path / "failed-attempt"

    def failing_writer(path: Path, value: bytes) -> None:
        if path.name == "resource_load.csv":
            raise OSError("injected")
        path.write_bytes(value)

    with pytest.raises(StandardExportError):
        write_change_report_export_package(
            package,
            failed,
            file_writer=failing_writer,
        )
    assert not failed.exists()
    assert not list(tmp_path.glob(".failed-attempt.tmp-*"))
